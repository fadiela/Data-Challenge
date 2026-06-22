import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import re
from pathlib import Path

# =========================
# KONFIGURASI
# =========================
st.set_page_config(
    page_title="Norm Database — Deka Insight",
    page_icon="📊",
    layout="wide"
)

RAW_DATA_PATH = Path(__file__).parent / "Deka_Insight_Data_For_Norm_Database.xlsx"

FILTER_COLS = [
    "Category", "Sub-Category", "Detail Product", "Gender",
    "Actual Age", "SES", "Occupation", "Type of Study",
    "Test Type", "Methodology", "Sub-Method", "# of Product", "Sequence"
]

DEMO_COLS = ["SbjNum", "No Project", "Category", "Sub-Category",
             "Detail Product", "Gender", "Actual Age", "SES",
             "Occupation", "Type of Study", "Test Type",
             "Methodology", "Sub-Method", "# of Product", "Sequence"]

# =========================
# LOAD RAW DATA
# =========================
@st.cache_data(show_spinner="Memuat raw data dari Excel...")
def load_raw_data(path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_dfs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        header_idx = next((i for i, r in enumerate(rows)
                           if r and str(r[0]).strip() == "SbjNum"), None)
        if header_idx is None:
            continue
        header = rows[header_idx]
        data_rows = rows[header_idx + 1:]
        if not data_rows:
            continue
        df = pd.DataFrame(data_rows, columns=header)
        valid_cols = [c for c in df.columns if c is not None]
        df = df[valid_cols]
        param_cols = [c for c in valid_cols if c not in DEMO_COLS]
        id_cols = [c for c in DEMO_COLS if c in valid_cols]
        df_long = df.melt(id_vars=id_cols, value_vars=param_cols,
                          var_name="raw_param", value_name="score")

        def parse_param(s):
            s = str(s).strip()
            m = re.search(r'-\s*(\d+)\s*pts?\s*$', s, re.IGNORECASE)
            if m:
                return s[:m.start()].strip().strip('-').strip(), int(m.group(1))
            return s, None

        df_long[["parameter_name", "scale_value"]] = df_long["raw_param"].apply(
            lambda x: pd.Series(parse_param(x))
        )
        df_long["score"] = pd.to_numeric(df_long["score"], errors="coerce")
        df_long = df_long.dropna(subset=["score", "scale_value"])
        df_long = df_long[df_long["score"] > 0]
        df_long["scale_value"] = df_long["scale_value"].astype(int)
        df_long["sheet"] = sheet_name
        all_dfs.append(df_long)

    if not all_dfs:
        return pd.DataFrame()
    combined = pd.concat(all_dfs, ignore_index=True)
    combined["parameter_name"] = combined["parameter_name"].str.strip()
    for col in FILTER_COLS:
        if col not in combined.columns:
            combined[col] = None
    return combined


# =========================
# FUNGSI HITUNG NORM
# =========================
def compute_norm_metrics(scores: pd.Series, scale: int) -> dict:
    n = len(scores)
    if n == 0:
        return {"TB%": None, "T2B%": None, "T3B%": None, "MS": None, "N": 0}
    tb  = (scores == scale).sum() / n * 100
    t2b = (scores >= scale - 1).sum() / n * 100
    t3b = (scores >= scale - 2).sum() / n * 100 if scale >= 7 else None
    ms  = scores.mean()
    return {
        "TB%":  round(tb, 1),
        "T2B%": round(t2b, 1),
        "T3B%": round(t3b, 1) if t3b is not None else None,
        "MS":   round(ms, 2),
        "N":    n
    }


def compute_norm_by_grade(scores: pd.Series, scale: int) -> dict:
    if len(scores) == 0:
        empty = {"TB%": None, "T2B%": None, "T3B%": None, "MS": None, "N": 0}
        return {"Top 25%": empty, "Average 50%": empty, "Bottom 25%": empty}
    n = len(scores)
    ranks = scores.rank(method="first")
    return {
        "Top 25%":     compute_norm_metrics(scores[ranks > n * 0.75], scale),
        "Average 50%": compute_norm_metrics(scores[(ranks > n * 0.25) & (ranks <= n * 0.75)], scale),
        "Bottom 25%":  compute_norm_metrics(scores[ranks <= n * 0.25], scale),
    }


def fmt_val(val, metric):
    if val is None:
        return "—"
    return f"{val:.2f}" if metric == "MS" else f"{val:.1f}%"


def get_verdict(prod_val, norm: dict, metric: str):
    top_val = norm["Top 25%"].get(metric)
    avg_val = norm["Average 50%"].get(metric)
    if prod_val is None or top_val is None or avg_val is None:
        return None, "#888", "—"
    if prod_val >= top_val:
        return "Above Average", "#16a34a", f"≥ Top 25% norm ({fmt_val(top_val, metric)})"
    elif prod_val >= avg_val:
        return "Average", "#d97706", f"Antara Average 50% ({fmt_val(avg_val, metric)}) dan Top 25% ({fmt_val(top_val, metric)})"
    else:
        return "Below Average", "#dc2626", f"< Average 50% norm ({fmt_val(avg_val, metric)})"


# =========================
# LOAD
# =========================
if not Path(RAW_DATA_PATH).exists():
    st.error(f"File tidak ditemukan: **{RAW_DATA_PATH}**")
    st.stop()

df = load_raw_data(str(RAW_DATA_PATH))
if df.empty:
    st.error("Data kosong.")
    st.stop()

all_sheets = sorted(df["sheet"].unique())

# =========================
# SIDEBAR
# =========================
with st.sidebar:
    st.header("📊 Norm Database")

    # --- Produk ---
    st.subheader("🎯 Produk yang Dinilai")
    selected_product_sheet = st.selectbox(
        "Project / Sheet sebagai Produk",
        options=["(Tidak dibandingkan)"] + list(all_sheets),
        help="Pilih project yang ingin di-benchmark terhadap norm database"
    )

    # --- Parameter & Metric ---
    st.divider()
    st.subheader("📐 Parameter & Metric")

    param_options = sorted(df["parameter_name"].dropna().unique())
    selected_param = st.selectbox(
        "Parameter",
        param_options,
        index=next((i for i, p in enumerate(param_options)
                    if "overall liking" in p.lower()), 0)
    )

    scale_options = sorted(
        df[df["parameter_name"] == selected_param]["scale_value"].dropna().unique()
    )
    selected_scale = st.selectbox(
        "Skala", scale_options, format_func=lambda x: f"{x} pts"
    )

    metric_options = ["TB%", "T2B%", "T3B%", "MS"] if selected_scale >= 7 else ["TB%", "T2B%", "MS"]
    selected_metric = st.selectbox("Metric Utama", metric_options)

    # --- Filter Dimensi ---
    st.divider()
    hdr_col, btn_col = st.columns([2, 1])
    hdr_col.subheader("🗂️ Filter Norm DB")
    clear_clicked = btn_col.button("🗑️ Clear", use_container_width=True,
                                   help="Reset semua filter dimensi")
    if clear_clicked:
        for col in FILTER_COLS:
            if f"fdim_{col}" in st.session_state:
                st.session_state[f"fdim_{col}"] = []

    # Build norm base (exclude produk dulu untuk dapat pilihan filter yg relevan)
    norm_base_df = df[
        (df["parameter_name"] == selected_param) &
        (df["scale_value"] == selected_scale)
    ].copy()
    if selected_product_sheet != "(Tidak dibandingkan)":
        norm_base_df = norm_base_df[norm_base_df["sheet"] != selected_product_sheet]

    active_filters = {}
    for col in FILTER_COLS:
        if col not in norm_base_df.columns:
            continue
        unique_vals = sorted(norm_base_df[col].dropna().astype(str).unique())
        if not unique_vals:
            continue
        sel = st.multiselect(col, options=unique_vals, default=[],
                             key=f"fdim_{col}")
        if sel:
            active_filters[col] = sel

    # Terapkan filter
    for col, vals in active_filters.items():
        norm_base_df = norm_base_df[norm_base_df[col].astype(str).isin(vals)]

# =========================
# HITUNG NORM & PRODUK
# =========================
norm_results = compute_norm_by_grade(norm_base_df["score"], selected_scale)

product_metrics = None
product_n = 0
if selected_product_sheet != "(Tidak dibandingkan)":
    prod_df = df[
        (df["sheet"] == selected_product_sheet) &
        (df["parameter_name"] == selected_param) &
        (df["scale_value"] == selected_scale)
    ]
    product_n = len(prod_df)
    if product_n > 0:
        product_metrics = compute_norm_metrics(prod_df["score"], selected_scale)

norm_n = len(norm_base_df)

# =========================
# HEADER
# =========================
st.title("📊 Norm Database Dashboard")
st.caption("Real-time norm benchmark · Deka Insight")

if active_filters:
    tags = " · ".join(f"**{k}**: {', '.join(v)}" for k, v in active_filters.items())
    st.info(f"🔍 Norm DB difilter: {tags}")

if norm_n == 0:
    st.warning("Tidak ada data norm untuk kombinasi filter ini.")
    st.stop()

# =========================
# SECTION 1 — KARTU RINGKAS
# =========================
prod_val = product_metrics.get(selected_metric) if product_metrics else None
top_val  = norm_results["Top 25%"].get(selected_metric)
avg_val  = norm_results["Average 50%"].get(selected_metric)
bot_val  = norm_results["Bottom 25%"].get(selected_metric)

verdict_label, verdict_color, verdict_desc = get_verdict(prod_val, norm_results, selected_metric)

c1, c2, c3, c4, c5 = st.columns(5)

def metric_card(col, title, value, subtitle="", bg="#f1f5f9", fg="#1e293b"):
    col.markdown(
        f"""<div style="background:{bg};border-radius:10px;padding:14px 16px;
                        text-align:center;height:100px;display:flex;flex-direction:column;
                        justify-content:center;">
            <div style="font-size:11px;color:{fg};opacity:0.7;margin-bottom:4px;">{title}</div>
            <div style="font-size:26px;font-weight:700;color:{fg};">{value}</div>
            <div style="font-size:10px;color:{fg};opacity:0.6;margin-top:2px;">{subtitle}</div>
        </div>""",
        unsafe_allow_html=True
    )

metric_card(c1, f"⬇️ Bottom 25% Norm", fmt_val(bot_val, selected_metric),
            f"N={norm_results['Bottom 25%']['N']:,}", "#fee2e2", "#991b1b")
metric_card(c2, f"📊 Average 50% Norm", fmt_val(avg_val, selected_metric),
            f"N={norm_results['Average 50%']['N']:,}", "#fef9c3", "#92400e")
metric_card(c3, f"⬆️ Top 25% Norm", fmt_val(top_val, selected_metric),
            f"N={norm_results['Top 25%']['N']:,}", "#dcfce7", "#166534")

if product_metrics and prod_val is not None:
    metric_card(c4, f"🎯 {selected_product_sheet}", fmt_val(prod_val, selected_metric),
                f"N={product_n:,}", "#eff6ff", "#1e40af")
    c5.markdown(
        f"""<div style="background:{verdict_color};border-radius:10px;padding:14px 16px;
                        text-align:center;height:100px;display:flex;flex-direction:column;
                        justify-content:center;">
            <div style="font-size:11px;color:white;opacity:0.85;margin-bottom:4px;">VERDICT</div>
            <div style="font-size:18px;font-weight:700;color:white;">{verdict_label}</div>
            <div style="font-size:10px;color:white;opacity:0.8;margin-top:2px;">{verdict_desc}</div>
        </div>""",
        unsafe_allow_html=True
    )
else:
    metric_card(c4, "🎯 Produk", "—", "Belum dipilih", "#f8fafc", "#94a3b8")
    metric_card(c5, "VERDICT", "—", "", "#f8fafc", "#94a3b8")

st.markdown("<br>", unsafe_allow_html=True)

# =========================
# SECTION 2 — GAUGE + TABEL
# =========================
gauge_col, table_col = st.columns([3, 2])

with gauge_col:
    st.markdown(f"#### 📏 Posisi Skor vs Norm — *{selected_param}* ({selected_scale} pts · {selected_metric})")

    scale_min = 1.0 if selected_metric == "MS" else 0.0
    scale_max = float(selected_scale) if selected_metric == "MS" else 100.0

    def to_pct(val):
        if val is None: return None
        return max(0.0, min(100.0, (val - scale_min) / (scale_max - scale_min) * 100))

    bot_pct  = to_pct(bot_val)  or 0
    avg_pct  = to_pct(avg_val)  or 0
    top_pct  = to_pct(top_val)  or 0
    prod_pct = to_pct(prod_val)

    w_bot  = bot_pct
    w_avg  = max(0, avg_pct - bot_pct)
    w_top  = max(0, top_pct - avg_pct)
    w_best = max(0, 100 - top_pct)

    prod_marker = ""
    if prod_pct is not None:
        prod_marker = f"""
        <div style="position:absolute;left:{prod_pct:.1f}%;top:-14px;
                    transform:translateX(-50%);z-index:10;pointer-events:none;">
            <div style="width:2px;height:58px;background:#f59e0b;margin:0 auto;"></div>
            <div style="background:#f59e0b;color:white;font-size:11px;font-weight:700;
                        border-radius:6px;padding:2px 7px;white-space:nowrap;
                        position:absolute;top:-22px;left:50%;transform:translateX(-50%);">
                {fmt_val(prod_val, selected_metric)}
            </div>
        </div>"""

    gauge_html = f"""
    <div style="margin:20px 0 56px 0;position:relative;">
        <div style="display:flex;justify-content:space-between;
                    font-size:11px;color:#64748b;margin-bottom:6px;font-weight:600;">
            <span>{fmt_val(scale_min, selected_metric)}</span>
            <span style="color:#dc2626;">Bot 25%: {fmt_val(bot_val, selected_metric)}</span>
            <span style="color:#d97706;">Avg 50%: {fmt_val(avg_val, selected_metric)}</span>
            <span style="color:#16a34a;">Top 25%: {fmt_val(top_val, selected_metric)}</span>
            <span>{fmt_val(scale_max, selected_metric)}</span>
        </div>
        <div style="position:relative;height:32px;border-radius:10px;
                    overflow:visible;display:flex;box-shadow:0 1px 4px rgba(0,0,0,.1);">
            <div style="width:{w_bot:.1f}%;background:#fca5a5;border-radius:10px 0 0 10px;"></div>
            <div style="width:{w_avg:.1f}%;background:#fde68a;"></div>
            <div style="width:{w_top:.1f}%;background:#86efac;"></div>
            <div style="width:{w_best:.1f}%;background:#16a34a;border-radius:0 10px 10px 0;"></div>
            {prod_marker}
        </div>
        <div style="display:flex;font-size:10px;color:#64748b;margin-top:6px;">
            <div style="width:{w_bot:.1f}%;text-align:center;">Below Avg</div>
            <div style="width:{w_avg:.1f}%;text-align:center;">Average</div>
            <div style="width:{w_top:.1f}%;text-align:center;">Good</div>
            <div style="width:{w_best:.1f}%;text-align:center;color:#16a34a;font-weight:600;">Top 25%</div>
        </div>
    </div>
    """
    st.markdown(gauge_html, unsafe_allow_html=True)

    # Bar chart semua metric
    if product_metrics:
        st.markdown(f"#### 📊 Semua Metric — Produk vs Norm")
        chart_rows = []
        for m in [x for x in ["TB%", "T2B%", "T3B%", "MS"] if x in metric_options]:
            pv = product_metrics.get(m)
            for grade, label in [("Bottom 25%", "Norm Bot 25%"),
                                  ("Average 50%", "Norm Avg 50%"),
                                  ("Top 25%",    "Norm Top 25%")]:
                nv = norm_results[grade].get(m)
                if nv is not None:
                    chart_rows.append({"Metric": m, "Sumber": label, "Nilai": nv})
            if pv is not None:
                chart_rows.append({"Metric": m, "Sumber": f"🎯 {selected_product_sheet}", "Nilai": pv})

        if chart_rows:
            pivot = (pd.DataFrame(chart_rows)
                     .pivot_table(index="Metric", columns="Sumber", values="Nilai"))
            # Reorder columns: produk di akhir
            prod_col = [c for c in pivot.columns if "🎯" in c]
            norm_cols = [c for c in pivot.columns if "🎯" not in c]
            pivot = pivot[norm_cols + prod_col]
            st.bar_chart(pivot, height=260)

with table_col:
    st.markdown("#### 📋 Tabel Perbandingan")

    rows_tbl = []
    grade_cfg = [
        ("Top 25%",    "⬆️ Top 25%",    "#dcfce7"),
        ("Average 50%","📊 Average 50%", "#fef9c3"),
        ("Bottom 25%", "⬇️ Bottom 25%", "#fee2e2"),
    ]
    for grade_key, grade_label, _ in grade_cfg:
        d = norm_results[grade_key]
        rows_tbl.append({
            "": grade_label,
            "N": f"{d['N']:,}",
            "TB%":  fmt_val(d["TB%"], "TB%"),
            "T2B%": fmt_val(d["T2B%"], "T2B%"),
            "T3B%": fmt_val(d["T3B%"], "T3B%") if selected_scale >= 7 else "—",
            "MS":   fmt_val(d["MS"], "MS"),
        })

    if product_metrics:
        rows_tbl.append({
            "": f"🎯 {selected_product_sheet}",
            "N": f"{product_n:,}",
            "TB%":  fmt_val(product_metrics.get("TB%"), "TB%"),
            "T2B%": fmt_val(product_metrics.get("T2B%"), "T2B%"),
            "T3B%": fmt_val(product_metrics.get("T3B%"), "T3B%") if selected_scale >= 7 else "—",
            "MS":   fmt_val(product_metrics.get("MS"), "MS"),
        })

    tbl_df = pd.DataFrame(rows_tbl)

    # Highlight kolom metric utama
    def highlight_metric_col(s):
        return ["font-weight:700;background:#e0f2fe" if s.name == selected_metric else ""
                for _ in s]

    styled_tbl = tbl_df.style.apply(highlight_metric_col, axis=0)
    st.dataframe(styled_tbl, use_container_width=True, hide_index=True)

    if product_metrics and verdict_label:
        st.markdown(
            f"""<div style="background:{verdict_color};color:white;border-radius:8px;
                            padding:10px 14px;text-align:center;margin-top:8px;">
                <span style="font-size:13px;font-weight:700;">VERDICT: {verdict_label}</span><br>
                <span style="font-size:11px;opacity:0.9;">{verdict_desc}</span>
            </div>""",
            unsafe_allow_html=True
        )

st.divider()

# =========================
# SECTION 3 — RINGKASAN SEMUA PARAMETER
# =========================
st.subheader(f"📑 Ringkasan Semua Parameter — Skala {selected_scale} pts")

if selected_product_sheet != "(Tidak dibandingkan)":
    st.caption(f"Metric utama: **{selected_metric}** · Produk: **{selected_product_sheet}** · "
               f"Warna verdict: 🟢 Above Average · 🟡 Average · 🔴 Below Average")
else:
    st.caption(f"Metric utama: **{selected_metric}** · Pilih produk di sidebar untuk melihat verdict per parameter.")

all_params_for_scale = sorted(
    df[df["scale_value"] == selected_scale]["parameter_name"].dropna().unique()
)

summary_rows = []
for param in all_params_for_scale:
    p_norm_df = df[
        (df["parameter_name"] == param) &
        (df["scale_value"] == selected_scale)
    ].copy()
    if selected_product_sheet != "(Tidak dibandingkan)":
        p_norm_df = p_norm_df[p_norm_df["sheet"] != selected_product_sheet]
    for col, vals in active_filters.items():
        if col in p_norm_df.columns:
            p_norm_df = p_norm_df[p_norm_df[col].astype(str).isin(vals)]
    if len(p_norm_df) == 0:
        continue

    p_norm = compute_norm_by_grade(p_norm_df["score"], selected_scale)
    row = {
        "Parameter":          param,
        "N Norm":             len(p_norm_df),
        "Bot 25%":            fmt_val(p_norm["Bottom 25%"].get(selected_metric), selected_metric),
        "Avg 50%":            fmt_val(p_norm["Average 50%"].get(selected_metric), selected_metric),
        "Top 25%":            fmt_val(p_norm["Top 25%"].get(selected_metric), selected_metric),
    }

    if selected_product_sheet != "(Tidak dibandingkan)":
        p_prod_df = df[
            (df["sheet"] == selected_product_sheet) &
            (df["parameter_name"] == param) &
            (df["scale_value"] == selected_scale)
        ]
        if len(p_prod_df) > 0:
            p_prod_m = compute_norm_metrics(p_prod_df["score"], selected_scale)
            p_val = p_prod_m.get(selected_metric)
            v_label, _, _ = get_verdict(p_val, p_norm, selected_metric)
            row["Produk"] = fmt_val(p_val, selected_metric)
            row["Verdict"] = v_label if v_label else "—"
        else:
            row["Produk"]  = "—"
            row["Verdict"] = "—"

    summary_rows.append(row)

if summary_rows:
    summary_df = pd.DataFrame(summary_rows)

    # Highlight baris produk yang sedang di-focus (selected_param)
    def highlight_selected_row(row):
        if row["Parameter"] == selected_param:
            return ["background:#dbeafe;font-weight:600"] * len(row)
        return [""] * len(row)

    def color_verdict_cell(val):
        return {
            "Above Average": "background:#dcfce7;color:#166534;font-weight:700",
            "Average":       "background:#fef9c3;color:#92400e;font-weight:700",
            "Below Average": "background:#fee2e2;color:#991b1b;font-weight:700",
        }.get(val, "")

    styled = summary_df.style.apply(highlight_selected_row, axis=1)
    if "Verdict" in summary_df.columns:
        styled = styled.map(color_verdict_cell, subset=["Verdict"])

    st.dataframe(styled, use_container_width=True, hide_index=True, height=420)

st.divider()

# =========================
# SECTION 4 — DISTRIBUSI & RAW DATA
# =========================
with st.expander("📊 Distribusi Skor — Norm DB vs Produk"):
    dist1, dist2 = st.columns(2)
    with dist1:
        st.markdown("**Distribusi Norm Database**")
        sc = (norm_base_df["score"].value_counts().sort_index()
              .reset_index().rename(columns={"score": "Skor", "count": "Frekuensi"}))
        sc["Skor"] = sc["Skor"].astype(int).astype(str)
        st.bar_chart(sc.set_index("Skor"))

    if product_metrics and selected_product_sheet != "(Tidak dibandingkan)":
        with dist2:
            st.markdown(f"**Distribusi Produk: {selected_product_sheet}**")
            prod_dist = df[
                (df["sheet"] == selected_product_sheet) &
                (df["parameter_name"] == selected_param) &
                (df["scale_value"] == selected_scale)
            ]
            sc2 = (prod_dist["score"].value_counts().sort_index()
                   .reset_index().rename(columns={"score": "Skor", "count": "Frekuensi"}))
            sc2["Skor"] = sc2["Skor"].astype(int).astype(str)
            st.bar_chart(sc2.set_index("Skor"))

with st.expander("🔍 Data Mentah Norm DB (200 baris pertama)"):
    show_cols = ["sheet", "parameter_name", "scale_value", "score"] + [
        c for c in FILTER_COLS if c in norm_base_df.columns
    ]
    st.dataframe(norm_base_df[show_cols].head(200), use_container_width=True)

st.divider()
csv_export = norm_base_df.to_csv(index=False).encode("utf-8")
st.download_button(
    "⬇️ Download data norm terfilter (CSV)",
    data=csv_export,
    file_name=f"norm_{selected_param}_{selected_scale}pts.csv",
    mime="text/csv"
)
